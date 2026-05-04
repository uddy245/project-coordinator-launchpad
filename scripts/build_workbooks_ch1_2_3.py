"""
Build XLSX starter workbooks for chapters 1, 2, 3.
Saves to public/templates/. Run once; commit the output to git.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT      = Path(__file__).resolve().parents[1]
TEMPLATES = ROOT / "public" / "templates"
TEMPLATES.mkdir(parents=True, exist_ok=True)

# ── Styles (university × office)
INK     = "FF11161E"
AMBER   = "FFB8801A"
RULE    = "FFD9DDE4"
PAPER   = "FFFFFFFF"
TINT    = "FFF5F7FA"

def header_style(cell):
    cell.font = Font(name="IBM Plex Sans", size=10, bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = Border(bottom=Side(style="thin", color=RULE))

def kicker_style(cell):
    cell.font = Font(name="IBM Plex Mono", size=9, color="FF6A7385", bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def title_style(cell):
    cell.font = Font(name="IBM Plex Serif", size=18, bold=True, color=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def subtitle_style(cell):
    cell.font = Font(name="IBM Plex Sans", size=11, color="FF2C3340")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def body_style(cell):
    cell.font = Font(name="IBM Plex Sans", size=10, color=INK)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

def hint_style(cell):
    cell.font = Font(name="IBM Plex Sans", size=9, italic=True, color="FF6A7385")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

def accent_pill(cell):
    cell.font = Font(name="IBM Plex Mono", size=9, bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=AMBER)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def thin_rule(ws, row, cols=8):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = Border(bottom=Side(style="thin", color=RULE))

def widen(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_header(ws, row=1):
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

# ────────────────────────────────────────────────────────────────────────────
# CHAPTER 1 — Coordinator Self-Audit
# ────────────────────────────────────────────────────────────────────────────

def build_ch1():
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Five Buckets Self-Score
    ws = wb.create_sheet("Five Buckets")
    widen(ws, [4, 22, 50, 14, 14, 38])

    ws.cell(row=2, column=2, value="MODULE 01 · SELF-AUDIT")
    kicker_style(ws.cell(row=2, column=2))
    ws.cell(row=3, column=2, value="The Five Buckets · Self-Score")
    title_style(ws.cell(row=3, column=2))
    ws.cell(row=4, column=2, value="Score yourself 1–5 on each bucket. Then pick ONE to work on for the next month. Don't tell anyone.")
    subtitle_style(ws.cell(row=4, column=2))
    ws.row_dimensions[3].height = 28

    headers = ["#", "Bucket", "What it means", "Score 1–5", "Trend ↑/→/↓", "Evidence / examples"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=i, value=h)
        header_style(c)
    ws.row_dimensions[6].height = 24

    buckets = [
        ("01", "Reliability",
         "If you say you'll do something, it gets done — on time, to standard."),
        ("02", "Signal clarity",
         "When you communicate, people know what's true, what matters, what to do."),
        ("03", "Pattern recognition",
         "You notice things before they blow up — slips, drift, friction."),
        ("04", "Judgment under ambiguity",
         "You act well when instructions are incomplete. You don't wait."),
        ("05", "Cumulative trust",
         "Built by every small thing done well over time. No shortcuts."),
    ]
    r = 7
    for code, name, meaning in buckets:
        ws.cell(row=r, column=1, value=code); body_style(ws.cell(row=r, column=1))
        ws.cell(row=r, column=2, value=name); body_style(ws.cell(row=r, column=2))
        ws.cell(row=r, column=2).font = Font(name="IBM Plex Sans", size=10, bold=True, color=INK)
        ws.cell(row=r, column=3, value=meaning); body_style(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value="");      body_style(ws.cell(row=r, column=4))
        ws.cell(row=r, column=5, value="");      body_style(ws.cell(row=r, column=5))
        ws.cell(row=r, column=6, value="");      body_style(ws.cell(row=r, column=6))
        ws.row_dimensions[r].height = 48
        thin_rule(ws, r, cols=6)
        r += 1

    r += 2
    ws.cell(row=r, column=2, value="Pick one bucket. Just one. Work it for 30 days.")
    ws.cell(row=r, column=2).font = Font(name="IBM Plex Serif", size=14, bold=True, color=INK)
    r += 1
    ws.cell(row=r, column=2, value="My pick:"); kicker_style(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value="________________________________"); body_style(ws.cell(row=r, column=3))
    r += 1
    ws.cell(row=r, column=2, value="Why:"); kicker_style(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value="________________________________"); body_style(ws.cell(row=r, column=3))

    freeze_header(ws, row=6)

    # ── Sheet 2: Inbox Audit
    ws = wb.create_sheet("Inbox Audit")
    widen(ws, [4, 30, 30, 16, 14, 30])
    ws.cell(row=2, column=2, value="MODULE 01 · TUESDAY MORNING DRILL")
    kicker_style(ws.cell(row=2, column=2))
    ws.cell(row=3, column=2, value="15-Minute Inbox Audit")
    title_style(ws.cell(row=3, column=2))
    ws.cell(row=4, column=2, value="Don't answer — audit. Look at the pattern of who is emailing you, what they want, how fast you reply.")
    subtitle_style(ws.cell(row=4, column=2))
    ws.row_dimensions[3].height = 28

    headers = ["#", "From (name / role)", "What they're asking for", "Action implied?", "Your response time", "What this reveals"]
    for i, h in enumerate(headers, start=1):
        header_style(ws.cell(row=6, column=i, value=h))
    ws.row_dimensions[6].height = 24

    for i in range(1, 11):
        ws.cell(row=6 + i, column=1, value=f"{i:02d}")
        body_style(ws.cell(row=6 + i, column=1))
        for c in range(2, 7):
            body_style(ws.cell(row=6 + i, column=c, value=""))
        ws.row_dimensions[6 + i].height = 28
        thin_rule(ws, 6 + i, cols=6)

    r = 19
    ws.cell(row=r, column=2, value="Pattern observation")
    ws.cell(row=r, column=2).font = Font(name="IBM Plex Serif", size=14, bold=True, color=INK)
    r += 1
    ws.cell(row=r, column=2, value="The information load comes mostly from:"); kicker_style(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value=""); body_style(ws.cell(row=r, column=3))
    r += 1
    ws.cell(row=r, column=2, value="My slowest response is to:"); kicker_style(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value=""); body_style(ws.cell(row=r, column=3))
    r += 1
    ws.cell(row=r, column=2, value="One change I'll make Monday:"); kicker_style(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3, value=""); body_style(ws.cell(row=r, column=3))

    freeze_header(ws, row=6)

    # ── Sheet 3: Role Clarity
    ws = wb.create_sheet("Role Clarity")
    widen(ws, [4, 38, 38, 38])
    ws.cell(row=2, column=2, value="MODULE 01 · WHICH ROLE ARE YOU IN")
    kicker_style(ws.cell(row=2, column=2))
    ws.cell(row=3, column=2, value="Coordinator Role — Self-Placement")
    title_style(ws.cell(row=3, column=2))
    ws.cell(row=4, column=2, value="Three coordinator roles cover the same job title. Knowing yours matters more than the title on LinkedIn.")
    subtitle_style(ws.cell(row=4, column=2))
    ws.row_dimensions[3].height = 28

    headers = ["Role", "What it covers", "Ceiling", "Are you in this one? (notes)"]
    for i, h in enumerate(headers, start=1):
        header_style(ws.cell(row=6, column=i, value=h))
    ws.row_dimensions[6].height = 24

    rows = [
        ("Administrative", "Logistics, scheduling, document libraries, attendance, formatting.", "Senior coordinator / project administrator"),
        ("Hybrid analyst", "All admin work + status reports, RAID log, change requests, dashboards, working-level meetings.", "Project manager"),
        ("Aspiring-PM", "In practice running a workstream with a coordinator title.", "Project manager (faster track)"),
    ]
    for i, (name, what, ceiling) in enumerate(rows):
        r = 7 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=1).font = Font(name="IBM Plex Sans", size=10, bold=True, color=INK)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        body_style(ws.cell(row=r, column=2, value=what))
        body_style(ws.cell(row=r, column=3, value=ceiling))
        body_style(ws.cell(row=r, column=4, value=""))
        ws.row_dimensions[r].height = 60
        thin_rule(ws, r, cols=4)

    r = 12
    ws.cell(row=r, column=2, value="The role I'm actually in:")
    ws.cell(row=r, column=2).font = Font(name="IBM Plex Sans", size=11, bold=True, color=INK)
    ws.cell(row=r, column=3, value=""); body_style(ws.cell(row=r, column=3))
    r += 1
    ws.cell(row=r, column=2, value="The role I want a year from now:")
    ws.cell(row=r, column=2).font = Font(name="IBM Plex Sans", size=11, bold=True, color=INK)
    ws.cell(row=r, column=3, value=""); body_style(ws.cell(row=r, column=3))
    r += 1
    ws.cell(row=r, column=2, value="Smallest move that shifts me toward it:")
    ws.cell(row=r, column=2).font = Font(name="IBM Plex Sans", size=11, bold=True, color=INK)
    ws.cell(row=r, column=3, value=""); body_style(ws.cell(row=r, column=3))

    out = TEMPLATES / "coordinator_role_self_audit.xlsx"
    wb.save(out)
    print(f"  ✓ {out.name}")


# ────────────────────────────────────────────────────────────────────────────
# CHAPTER 2 — Lifecycle Support Map
# ────────────────────────────────────────────────────────────────────────────

def build_ch2():
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Phase Map
    ws = wb.create_sheet("Phase Map")
    widen(ws, [4, 16, 38, 24, 28, 26])
    ws.cell(row=2, column=2, value="MODULE 02 · LIFECYCLE SUPPORT MAP")
    kicker_style(ws.cell(row=2, column=2))
    ws.cell(row=3, column=2, value="One Project, Five Phases")
    title_style(ws.cell(row=3, column=2))
    ws.cell(row=4, column=2, value="For one of your real projects, fill in what the coordinator does at each phase. Use this on Day 1 of any new assignment.")
    subtitle_style(ws.cell(row=4, column=2))
    ws.row_dimensions[3].height = 28

    headers = ["Phase", "Coordinator's primary task", "Key artifact you'd maintain", "Next gate / milestone", "AI / info-mgmt guardrail"]
    for i, h in enumerate(headers, start=2):
        header_style(ws.cell(row=6, column=i, value=h))
    ws.cell(row=6, column=1, value="#")
    header_style(ws.cell(row=6, column=1))
    ws.row_dimensions[6].height = 28

    phases = [
        ("01", "Initiation", "Gather essentials. Confirm approval state. Identify sponsors and stakeholders. Find source of truth.",
         "Charter / Stakeholder log",
         "First gate or governance touch",
         "Who can see the charter? Where does it live?"),
        ("02", "Planning", "Set up trackers, cadences, folders, naming conventions. Draft status template.",
         "Schedule / RAID log / Action tracker",
         "Plan baseline approval",
         "One source of truth — not multiple trackers."),
        ("03", "Execution", "Move actions to owners and dates. Keep issues visible. Vendor follow-up.",
         "Action log / Status report",
         "Major build milestone",
         "Don't paste sensitive vendor info into public AI tools."),
        ("04", "Monitoring", "Compare actual to plan. Surface decisions. Separate risks from issues.",
         "Variance log / Steering pack",
         "Steering committee decision",
         "Verify any AI summary against trusted sources."),
        ("05", "Closing", "Confirm action burn-down. Archive. Capture lessons. Hand over to operations.",
         "Closure report / Lessons learned",
         "Service readiness / Post-impl review",
         "Archive controlled versions, not screenshots."),
    ]
    for i, (code, phase, task, artifact, gate, ai) in enumerate(phases):
        r = 7 + i
        ws.cell(row=r, column=1, value=code); body_style(ws.cell(row=r, column=1))
        ws.cell(row=r, column=2, value=phase)
        ws.cell(row=r, column=2).font = Font(name="IBM Plex Sans", size=10, bold=True, color=INK)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="top", indent=1)
        body_style(ws.cell(row=r, column=3, value=task))
        body_style(ws.cell(row=r, column=4, value=artifact))
        body_style(ws.cell(row=r, column=5, value=gate))
        body_style(ws.cell(row=r, column=6, value=ai))
        ws.row_dimensions[r].height = 70
        thin_rule(ws, r, cols=6)

    freeze_header(ws, row=6)

    # ── Sheet 2: Phase Checklist
    ws = wb.create_sheet("Phase Checklist")
    widen(ws, [4, 8, 56, 36])
    ws.cell(row=2, column=2, value="MODULE 02 · SIX QUESTIONS")
    kicker_style(ws.cell(row=2, column=2))
    ws.cell(row=3, column=2, value="The Phase Boundary Checklist")
    title_style(ws.cell(row=3, column=2))
    ws.cell(row=4, column=2, value="Run these six questions at the start of every phase or major handoff.")
    subtitle_style(ws.cell(row=4, column=2))
    ws.row_dimensions[3].height = 28

    headers = ["✓", "Question", "Your answer for THIS phase"]
    for i, h in enumerate(headers, start=2):
        header_style(ws.cell(row=6, column=i, value=h))
    ws.cell(row=6, column=1, value="#"); header_style(ws.cell(row=6, column=1))
    ws.row_dimensions[6].height = 26

    questions = [
        "What phase are we in, and what is the next gate, milestone, or review?",
        "What evidence does the team need for that decision?",
        "Which documents are authoritative — and where is the source of truth?",
        "Which actions, issues, risks, and changes need updating now?",
        "Who needs an update, approval, or escalation before the next milestone?",
        "What must be handed forward so nothing gets lost?",
    ]
    for i, q in enumerate(questions):
        r = 7 + i
        ws.cell(row=r, column=1, value=f"{i+1:02d}"); body_style(ws.cell(row=r, column=1))
        ws.cell(row=r, column=2, value="☐"); body_style(ws.cell(row=r, column=2))
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="top")
        body_style(ws.cell(row=r, column=3, value=q))
        ws.cell(row=r, column=3).font = Font(name="IBM Plex Sans", size=10, bold=True, color=INK)
        body_style(ws.cell(row=r, column=4, value=""))
        ws.row_dimensions[r].height = 56
        thin_rule(ws, r, cols=4)

    freeze_header(ws, row=6)

    # ── Sheet 3: Role Boundaries
    ws = wb.create_sheet("Role Boundaries")
    widen(ws, [4, 40, 28, 28])
    ws.cell(row=2, column=2, value="MODULE 02 · WHO OWNS WHAT")
    kicker_style(ws.cell(row=2, column=2))
    ws.cell(row=3, column=2, value="Role Boundary Map")
    title_style(ws.cell(row=3, column=2))
    ws.cell(row=4, column=2, value="Coordination is not authority. Mark each scenario: who owns it, what's your support role.")
    subtitle_style(ws.cell(row=4, column=2))
    ws.row_dimensions[3].height = 28

    headers = ["Question / decision", "Who owns it", "Your supporting role"]
    for i, h in enumerate(headers, start=2):
        header_style(ws.cell(row=6, column=i, value=h))
    ws.cell(row=6, column=1, value="#"); header_style(ws.cell(row=6, column=1))
    ws.row_dimensions[6].height = 26

    cases = [
        "Approving the project, funding, or major direction change",
        "Day-to-day trade-offs between scope, schedule, and cost",
        "Backlog ordering and product value (Scrum)",
        "Removing impediments and Scrum effectiveness",
        "Logs, follow-up, status packaging, repository discipline",
        "A vendor missing a deliverable date — who escalates",
        "Whether a change request is approved",
        "Whether minutes go out today or tomorrow",
    ]
    for i, q in enumerate(cases):
        r = 7 + i
        ws.cell(row=r, column=1, value=f"{i+1:02d}"); body_style(ws.cell(row=r, column=1))
        body_style(ws.cell(row=r, column=2, value=q))
        body_style(ws.cell(row=r, column=3, value=""))
        body_style(ws.cell(row=r, column=4, value=""))
        ws.row_dimensions[r].height = 36
        thin_rule(ws, r, cols=4)

    freeze_header(ws, row=6)

    out = TEMPLATES / "lifecycle_support_map_starter.xlsx"
    wb.save(out)
    print(f"  ✓ {out.name}")


# ────────────────────────────────────────────────────────────────────────────
# CHAPTER 3 — Editing Pass Workbook
# ────────────────────────────────────────────────────────────────────────────

def build_ch3():
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Editing Pass Tracker
    ws = wb.create_sheet("Editing Pass")
    widen(ws, [4, 14, 50, 50, 12, 18])
    ws.cell(row=2, column=2, value="MODULE 03 · THE 90-SECOND HABIT")
    kicker_style(ws.cell(row=2, column=2))
    ws.cell(row=3, column=2, value="Editing Pass Tracker")
    title_style(ws.cell(row=3, column=2))
    ws.cell(row=4, column=2, value="For five emails of consequence this week, paste the draft, then the edited version. Track your cut percentage.")
    subtitle_style(ws.cell(row=4, column=2))
    ws.row_dimensions[3].height = 28

    headers = ["Date", "Recipient / register", "Draft (before)", "Edited (after)", "% cut", "Hedges removed"]
    for i, h in enumerate(headers, start=2):
        header_style(ws.cell(row=6, column=i, value=h))
    ws.cell(row=6, column=1, value="#"); header_style(ws.cell(row=6, column=1))
    ws.row_dimensions[6].height = 26

    for i in range(1, 6):
        r = 6 + i
        ws.cell(row=r, column=1, value=f"{i:02d}"); body_style(ws.cell(row=r, column=1))
        for c in range(2, 7):
            body_style(ws.cell(row=r, column=c, value=""))
        ws.row_dimensions[r].height = 90
        thin_rule(ws, r, cols=6)

    r = 14
    ws.cell(row=r, column=2, value="Pattern reflection")
    ws.cell(row=r, column=2).font = Font(name="IBM Plex Serif", size=14, bold=True, color=INK)
    r += 1
    ws.cell(row=r, column=2, value="My most common hedge:"); kicker_style(ws.cell(row=r, column=2))
    body_style(ws.cell(row=r, column=3, value=""))
    r += 1
    ws.cell(row=r, column=2, value="My most common padding phrase:"); kicker_style(ws.cell(row=r, column=2))
    body_style(ws.cell(row=r, column=3, value=""))
    r += 1
    ws.cell(row=r, column=2, value="Average % cut across the week:"); kicker_style(ws.cell(row=r, column=2))
    body_style(ws.cell(row=r, column=3, value=""))

    freeze_header(ws, row=6)

    # ── Sheet 2: Subject Line Audit
    ws = wb.create_sheet("Subject Lines")
    widen(ws, [4, 50, 50, 22])
    ws.cell(row=2, column=2, value="MODULE 03 · THE FIRST LEDE")
    kicker_style(ws.cell(row=2, column=2))
    ws.cell(row=3, column=2, value="Subject Line Audit")
    title_style(ws.cell(row=3, column=2))
    ws.cell(row=4, column=2, value="Pull your last 10 sent emails. Rate each subject line, then rewrite the weak ones.")
    subtitle_style(ws.cell(row=4, column=2))
    ws.row_dimensions[3].height = 28

    headers = ["Original subject line", "Improved version (lede + action)", "Verdict"]
    for i, h in enumerate(headers, start=2):
        header_style(ws.cell(row=6, column=i, value=h))
    ws.cell(row=6, column=1, value="#"); header_style(ws.cell(row=6, column=1))
    ws.row_dimensions[6].height = 26

    examples = [
        ("Update", "Vendor delivery slipped — needs decision on testing start"),
        ("Quick question", "Need your sign-off on CR-004 by Thursday"),
        ("Following up", "Status report attached — amber on integration"),
    ]
    for i, (orig, better) in enumerate(examples):
        r = 7 + i
        ws.cell(row=r, column=1, value=f"e{i+1}"); body_style(ws.cell(row=r, column=1))
        body_style(ws.cell(row=r, column=2, value=orig))
        ws.cell(row=r, column=2).font = Font(name="IBM Plex Mono", size=10, color="FF6A7385", italic=True)
        body_style(ws.cell(row=r, column=3, value=better))
        ws.cell(row=r, column=3).font = Font(name="IBM Plex Mono", size=10, color=INK)
        body_style(ws.cell(row=r, column=4, value="example"))
        ws.cell(row=r, column=4).font = Font(name="IBM Plex Sans", size=9, italic=True, color="FF6A7385")
        ws.row_dimensions[r].height = 36
        thin_rule(ws, r, cols=4)

    for i in range(10):
        r = 10 + i
        ws.cell(row=r, column=1, value=f"{i+1:02d}"); body_style(ws.cell(row=r, column=1))
        for c in range(2, 5):
            body_style(ws.cell(row=r, column=c, value=""))
        ws.row_dimensions[r].height = 32
        thin_rule(ws, r, cols=4)

    freeze_header(ws, row=6)

    # ── Sheet 3: Bad-News Restructure
    ws = wb.create_sheet("Bad News")
    widen(ws, [4, 22, 60, 30])
    ws.cell(row=2, column=2, value="MODULE 03 · DELIVERING BAD NEWS")
    kicker_style(ws.cell(row=2, column=2))
    ws.cell(row=3, column=2, value="Four-Sentence Restructure")
    title_style(ws.cell(row=3, column=2))
    ws.cell(row=4, column=2, value="Pick a piece of bad news you need to deliver this month. Compose it in four sentences.")
    subtitle_style(ws.cell(row=4, column=2))
    ws.row_dimensions[3].height = 28

    structure = [
        ("Sentence 01", "The bad news, plainly stated.",
         "What is the actual bad news in one sentence?"),
        ("Sentence 02", "Minimal context to understand it.",
         "What's the smallest amount of context the reader needs?"),
        ("Sentence 03", "What is being done.",
         "Who's working on it, by when, what changes?"),
        ("Sentence 04", "What you need from the reader.",
         "Decision? Acknowledgement? Briefing time? Nothing?"),
    ]
    headers = ["Step", "Purpose", "Your sentence"]
    for i, h in enumerate(headers, start=2):
        header_style(ws.cell(row=6, column=i, value=h))
    ws.cell(row=6, column=1, value="#"); header_style(ws.cell(row=6, column=1))
    ws.row_dimensions[6].height = 26

    for i, (step, purpose, hint) in enumerate(structure):
        r = 7 + i
        ws.cell(row=r, column=1, value=step); body_style(ws.cell(row=r, column=1))
        ws.cell(row=r, column=1).font = Font(name="IBM Plex Mono", size=10, bold=True, color=AMBER)
        body_style(ws.cell(row=r, column=2, value=purpose))
        ws.cell(row=r, column=2).font = Font(name="IBM Plex Sans", size=10, bold=True, color=INK)
        body_style(ws.cell(row=r, column=3, value=hint))
        ws.cell(row=r, column=3).font = Font(name="IBM Plex Sans", size=10, italic=True, color="FF6A7385")
        body_style(ws.cell(row=r, column=4, value=""))
        ws.row_dimensions[r].height = 64
        thin_rule(ws, r, cols=4)

    r = 13
    ws.cell(row=r, column=2, value="Now read it as one paragraph and edit again.")
    ws.cell(row=r, column=2).font = Font(name="IBM Plex Serif", size=14, bold=True, color=INK)
    r += 1
    body_style(ws.cell(row=r, column=2, value="Final version:"))
    ws.cell(row=r, column=2).font = Font(name="IBM Plex Sans", size=10, bold=True, color=INK)
    body_style(ws.cell(row=r, column=3, value=""))
    ws.row_dimensions[r].height = 80

    freeze_header(ws, row=6)

    out = TEMPLATES / "written_voice_editing_pass.xlsx"
    wb.save(out)
    print(f"  ✓ {out.name}")


# ────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"Writing workbook templates to {TEMPLATES}/")
    build_ch1()
    build_ch2()
    build_ch3()
    print("Done.")
